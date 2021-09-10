import os
import openpyxl

from openpyxl.workbook.workbook import Workbook

wb = openpyxl.load_workbook('/Users/techmate/Training/ViT/PythonDataScience/itp_week_3/day_2/lecture.xlsx')
# print(str(wb))
# print(wb.sheetnames)

# wb = Workbook()
# print(str(wb))

# new_sheet = wb.create_sheet()
for i in range(3,10):
    wb.create_sheet()

for sheet in wb:
    y = 50
    sheet.title = "MyNewSheet" + str(y)
    y += 1
    
# print(str(new_sheet))
print(str(wb.sheetnames))