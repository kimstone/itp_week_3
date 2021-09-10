import openpyxl

filename = "/Users/techmate/Training/ViT/PythonDataScience/itp_week_3/day_1/lecture.xlsx"

wb = openpyxl.load_workbook(filename)

sheet = wb['Sheet1']
# print(sheet)
row_cnt = sheet.max_row
print("Sheet Row Count: " + str(row_cnt))

# print(type(my_sheet))

#sheet_cell = my_sheet['B1']

date_cell = sheet['A1']
print(type(date_cell.value))

datetime = date_cell.value
print(datetime.date())

# for i in range(1, row_cnt + 1):
#     # on the date of A, C amount of B were sold
#     # worksheet['A'+ i]????
#     date = "A" + str(i)
#     date_cell = sheet[date]

#     amount = "C" + str(i)
#     amount_cell = sheet[amount]

#     fruit = "B" + str(i)
#     fruit_cell = sheet[fruit]

#     print("On the date of " + str(date_cell.value.date()) + ", " + str(amount_cell.value) + " amount of " + fruit_cell.value + "!")

for i in range(wb.sheetnames.__len__):
    sheet = wb[wb.sheetnames[i]]